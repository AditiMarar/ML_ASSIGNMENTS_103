import numpy as np
from collections import Counter
from openpyxl import load_workbook

def entropy(labels):#l is for labels
    lc=Counter(labels)#lc is label count
    e=0.0#entropy
    ts=len(labels)#total samples
    for l in lc:
        p=lc[l]/ts#calculating probability
        e-=p*np.log2(p)
    return e
def info_g(d,labels,fi):#d is data, fi is feature index
    te=entropy(labels)#total entropy
    unique_val=set(d[:,fi])
    ne=0.0#new entropy
    for v in unique_val:
        s_i=np.where(d[:,fi]==v)[0]# subset indices
        s_l=labels[s_i]#subset labels
        s_e=entropy(s_l)#subset entropy
        s_w=len(s_i)/len(labels)
        ne+=s_w*s_e
    info_gain=te-ne
    return info_gain
def find_root_feature(d,labels):
    nf=d.shape[1]
    best_f=None#best feature
    best_g=-1#best gain
    for fi in range(nf):
        g=info_g(d,labels,fi)#gain
        if g>best_g:#updating feature and gain
            best_g=g
            best_f=fi
    return best_f

file='d2.xlsx'
wb=load_workbook(file)
s=wb.active
l=[]
d=[]
for r in s.iter_rows(values_only=True):
    l+=[r[5]]
    d+=[[r[4],r[0],r[1]]]
data=np.array(d)
labels=np.array(l)
root_fi=find_root_feature(data,labels)
print(root_fi)
wb.close()
