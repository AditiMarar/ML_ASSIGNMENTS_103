import openpyxl
import numpy as np
import matplotlib.pyplot as plt
#the name of the file is different as i was unable to access it and was denied permission therefore using a different file
def get_a_c(data):
    A=[]
    C=[]
    for row in data.iter_rows(values_only=True):
        A.append(list(row[1:4]))
        C.append(list(row[4:5]))
    C.pop(0)
    head=A.pop(0)
    return A,C,head
def dim_vec(a,c):
    return len(c[0]),len(a[0]),len(a)
def rank_of_mat(a):
    mat = np.array(a)
    r= np.linalg.matrix_rank(mat)
    return r
def cost(a,c):
    A=np.array(a)
    C=np.array(c)
    A_pinv=np.linalg.pinv(A)#getting pseudoinverse of A
    product_cost= np.dot(A_pinv,C)
    return product_cost
def edit_status(c,data):
    r_p=[]#r_p will include the status, whether they are rich or poor
    
    for row in c:
        if row[0]>=200:
            r_p+=["RICH"]
        else:
            r_p+=["POOR"]
        
    column_index=1
    data.cell(row=1, column=column_index, value="Status")
    for i,status in enumerate(r_p,start=2):
        data.cell(row=i, column=column_index, value=status)
    return r_p

def mean_var(a):
    mean=sum(a)/len(a)
    var_l=[]
    for i in a:
        var_l+=[(i-mean)**2]
    var=sum(var_l)/len(var_l)
    return mean,var
def get_price_days_months_chg(data):
    price=[]
    days=[]
    months=[]
    chg=[]
    neg=[]
    for i in data['D'][1:]:
        price+=[i.value]
    for i in data['C'][1:]:
        days+=[i.value]
    for i in data['B'][1:]:
        months+=[i.value]
    for i in data['I'][1:]:
        chg+=[i.value]
        if i.value<0:
            neg+=[i.value]
    return price,days,months,chg,neg
def get_apr_wed(price,months,days):
    apr=[]
    for i in range(len(months)):
        if months[i]=="Apr":
            apr+=[price[i]]
    wed=[]
    for i in range(len(days)):
        if days[i]=="Wed":
            wed+=[price[i]]
    return apr,wed
def get_x(days):
    x=[]
    for i in days:
        if i=="Mon":
            x+=[1]
        elif i=="Tue":
            x+=[2]
        elif i=="Wed":
            x+=[3]
        elif i=="Thu":
            x+=[4]
        else:
            x+=[5]
    return x
def profit_wed(chg,days,wed):
    w_p=0
    for i in range(len(chg)):
        if chg[i]>0 and days[i]=="Wed":
            w_p+=1
    return w_p,len(chg),len(wed)

    
workbook = openpyxl.load_workbook("lab3.xlsx")
data_1 = workbook["Sheet1"]
print("\n\t\tLAB3 - MACHINE LEARNING AIE22103\n\n")
ac=get_a_c(data_1)
A=ac[0]
C=ac[1]
print("Matrix A is~")
for i in A:
    print(i)
print("Matrix C is~")
for i in C:
    print(i)
dim_n_vec=dim_vec(A,C)
print("Dimentionality of matrix C~",dim_n_vec[0])
print("Dimentionality of matrix A~",dim_n_vec[1])
print("Total no. of vectors in vector space is",dim_n_vec[2])
print("Rank of matrix A is~",rank_of_mat(A))
print("Cost of each item is~")
for i,c in enumerate(cost(A,C),start=0):
    print(ac[2][i],"costs",round(c[0]))
stat=edit_status(C,data_1)
c_no=1
for i in stat:
    print("Customer",c_no,"is~",i)
    c_no+=1
workbook.save("lab3.xlsx")
print("================================================================================================")

wb=openpyxl.load_workbook("lab3.xlsx")
data_2=wb["Sheet2"]
all_info=get_price_days_months_chg(data_2)
price=all_info[0]
days = all_info[1]
months= all_info[2]
chg =all_info[3]
neg=all_info[4]
a_w=get_apr_wed(price,months,days)
apr=a_w[0]
wed=a_w[1]
print("Mean :",mean_var(price)[0])
print("Variance :",mean_var(price)[1])
print("Mean of all wednesdays:",mean_var(wed)[0])
print("Variance of all wednesdays:",mean_var(wed)[1])
print("Mean difference=",mean_var(price)[0]-mean_var(wed)[0])
print("Variance difference=",mean_var(price)[1]-mean_var(wed)[1])
print("Mean of april :",mean_var(apr)[0])
print("Variance of april :",mean_var(apr)[1])
print("Mean difference:",mean_var(price)[0]-mean_var(apr)[0])
print("Variance difference:",mean_var(apr)[1]-mean_var(apr)[1])
print("Probability of loss=",len(neg)/len(chg))
p_w=profit_wed(chg,days,wed)
print("Probability of profit on wednesday=",p_w[0]/p_w[1])
print("Conditional probability of profit given that it is a wednesday=",p_w[0]/p_w[2])
x=get_x(days)
plt.scatter(x,chg)
plt.show()
